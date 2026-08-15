"""
ETL Crédito Externo (SCCI) — Bitácora 2026-I
Carga tres tablas:
  credito_portafolio          ← hoja "Portafolio" (17 créditos BID/BM/CAF, USD)
  credito_ejecucion_entidad   ← hoja "Ejecución Entidad" (recurso 13/14, COP → mmm)
  credito_ejecucion_historica ← hoja "Comp Ejecución Anual Marzo" (2023-2026, COP → mmm)
"""
import openpyxl, sys

import bases
import db as dbmod

FILE = bases.excel(8, "Datos informe*.xlsx")

def clean(v):
    return str(v).strip() if v is not None else ''

wb = openpyxl.load_workbook(FILE, data_only=True)

# ── 1. Portafolio: 17 créditos ───────────────────────────────────────────────
ws = wb['Portafolio']
rows = list(ws.iter_rows(values_only=True))
header = [clean(h) for h in rows[0]]
port_rows = []
for r in rows[1:]:
    if not r[0]:
        continue
    nombre, nombre_corto, fuente, contrato, monto, desembolsado, contratante, sector = r
    port_rows.append((nombre, nombre_corto, clean(fuente), str(contrato) if contrato is not None else None,
                       clean(sector), float(monto) if monto is not None else None,
                       float(desembolsado) if desembolsado is not None else 0.0))

print(f"Portafolio: {len(port_rows)} créditos leídos", file=sys.stderr)

# ── 2. Ejecución Entidad: recurso 13/14 ──────────────────────────────────────
ws2 = wb['Ejecución Entidad']
rows2 = list(ws2.iter_rows(values_only=True))
ent_rows = []
for r in rows2[1:]:
    if not r[0]:
        continue
    entidad, sector, apr_inicial, apr_vigente, compromiso, obligacion, pago, pct_com, pct_ejec, pct_pago = r
    ent_rows.append((
        clean(entidad), clean(sector),
        round(apr_inicial / 1e9, 3) if apr_inicial else None,
        round(apr_vigente / 1e9, 3) if apr_vigente else None,
        round(compromiso / 1e9, 3) if compromiso else None,
        round(obligacion / 1e9, 3) if obligacion else None,
        round(pago / 1e9, 3) if pago else None,
        round(pct_com * 100, 2) if pct_com is not None else None,
        round(pct_ejec * 100, 2) if pct_ejec is not None else None,
        round(pct_pago * 100, 2) if pct_pago is not None else None,
    ))

print(f"Ejecución Entidad: {len(ent_rows)} entidades leídas", file=sys.stderr)

# ── 3. Comparativo histórico anual (2023-2026) ───────────────────────────────
ws3 = wb['Comp Ejección Anual Marzo']  # nombre de hoja tal cual viene en el Excel (con typo "Ejección")
rows3 = list(ws3.iter_rows(values_only=True))
hist_rows = []
for r in rows3[1:]:
    if not r[0]:
        continue
    anio, comp_pct, ejec_pct, pag_pct, vigente, comp_val, ejec_val, pag_val = r[:8]
    hist_rows.append((
        int(anio),
        round(comp_pct * 100, 2) if comp_pct is not None else None,
        round(ejec_pct * 100, 2) if ejec_pct is not None else None,
        round(pag_pct * 100, 2) if pag_pct is not None else None,
        round(vigente / 1e9, 3) if vigente else None,
        round(comp_val / 1e9, 3) if comp_val else None,
        round(ejec_val / 1e9, 3) if ejec_val else None,
        round(pag_val / 1e9, 3) if pag_val else None,
    ))

print(f"Ejecución histórica: {len(hist_rows)} años leídos", file=sys.stderr)

wb.close()

# ── 4. Cargar en BD ───────────────────────────────────────────────────────────
conn = dbmod.conectar()

bid = dbmod.bitacora_reciente(conn)
print(f"bitacora_id={bid}", file=sys.stderr)

# Antes se hacía DROP + CREATE de las tres tablas desde schema.sql; ahora el
# esquema lo gobierna db/mssql/ y basta con vaciar esta bitácora.
# credito_portafolio no tiene clave natural única, así que se recarga entera.
conn.vaciar_bitacora(
    ("credito_portafolio", "credito_ejecucion_entidad", "credito_ejecucion_historica"),
    bid,
)

conn.executemany("""
    INSERT INTO dbo.credito_portafolio
        (bitacora_id, nombre, nombre_corto, fuente, contrato, sector, monto_usd, desembolsado_usd)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
""", [(bid, *r) for r in port_rows])

conn.upsert(
    "credito_ejecucion_entidad",
    ["bitacora_id", "entidad", "sector", "apr_inicial_mmm", "apr_vigente_mmm",
     "compromiso_mmm", "obligacion_mmm", "pago_mmm", "pct_com", "pct_ejec", "pct_pago"],
    [(bid, *r) for r in ent_rows], claves=["bitacora_id", "entidad"],
)

conn.upsert(
    "credito_ejecucion_historica",
    ["bitacora_id", "anio", "pct_comprometido", "pct_ejecutado", "pct_pagado",
     "vigente_mmm", "comprometido_mmm", "ejecutado_mmm", "pagado_mmm"],
    [(bid, *r) for r in hist_rows], claves=["bitacora_id", "anio"],
)

conn.commit()

n1 = conn.execute("SELECT COUNT(*) FROM credito_portafolio WHERE bitacora_id=?", (bid,)).fetchone()[0]
n2 = conn.execute("SELECT COUNT(*) FROM credito_ejecucion_entidad WHERE bitacora_id=?", (bid,)).fetchone()[0]
n3 = conn.execute("SELECT COUNT(*) FROM credito_ejecucion_historica WHERE bitacora_id=?", (bid,)).fetchone()[0]

tot = conn.execute("""
    SELECT COUNT(*), ROUND(SUM(monto_usd),2), ROUND(SUM(desembolsado_usd),2)
    FROM credito_portafolio WHERE bitacora_id=?
""", (bid,)).fetchone()

conn.close()

print(f"\n[OK] credito_portafolio:          {n1} filas cargadas")
print(f"[OK] credito_ejecucion_entidad:   {n2} filas cargadas")
print(f"[OK] credito_ejecucion_historica: {n3} filas cargadas")
print(f"\n=== VERIFICACIÓN ===")
print(f"  Cartera vigente: {tot[0]} operaciones")
print(f"  Total USD:       {tot[1]:,.2f}")
print(f"  Desembolsado:    {tot[2]:,.2f}")
