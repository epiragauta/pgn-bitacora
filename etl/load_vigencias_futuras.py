"""
ETL Vigencias Futuras — Bitácora 2026-I
Carga dos tablas:
  deflactores_pib   ← TD BITACORA (deflactor, PIB corriente, PIB constante)
  vigencias_futuras ← BASE_SIIF_2 pivot (sector × año) en pesos corrientes mmm
"""
import openpyxl, sys
from collections import defaultdict

import bases
import db as dbmod

FILE = bases.excel(5, "*Base VF*.xlsx")

def clean(v):
    return str(v).strip() if v is not None else ''

def to_f(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = clean(v).replace('\xa0', '').replace(' ', '').replace(',', '.')
    try: return float(s)
    except: return None

# ── 1. TD BITACORA: deflactores y PIB ────────────────────────────────────────
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
ws_td = wb["TD BITACORA"]

td_rows = []
for i, row in enumerate(ws_td.iter_rows(values_only=True)):
    td_rows.append(list(row))
    if i >= 75:
        break

# Encontrar fila cabecera de años ("Etiquetas de fila")
year_col = {}   # year -> col_index
for r in td_rows:
    if clean(r[0]).lower().startswith('etiquetas de fila'):
        for c, v in enumerate(r):
            try:
                y = int(clean(v))
                if 2025 <= y <= 2054:
                    year_col[y] = c
            except:
                pass
        break

if not year_col:
    sys.exit("ERROR: no se encontró cabecera de años en TD BITACORA")

print(f"Años encontrados: {sorted(year_col)}", file=sys.stderr)

deflactor   = {}   # year -> float
pib_corr    = {}   # year -> float (mmm corrientes)
pib_ctes    = {}   # year -> float (mmm constantes 2026)

for r in td_rows:
    lbl = clean(r[0]).upper()
    if 'DEFLACTOR PIB' in lbl and 'BASE' in lbl:
        for y, c in year_col.items():
            v = to_f(r[c])
            if v: deflactor[y] = v
    elif lbl.startswith('PIB CORRIENTES'):
        for y, c in year_col.items():
            # puede estar en millones de pesos → dividir entre 1000 para obtener mmm
            v = to_f(r[c])
            if v: pib_corr[y] = round(v / 1000, 3)   # millones → mmm
    elif lbl.startswith('PIB CONSTANTES'):
        for y, c in year_col.items():
            v = to_f(r[c])
            if v: pib_ctes[y] = v   # ya en mmm (miles de millones)

print(f"Deflactores encontrados: {len(deflactor)} años", file=sys.stderr)
print(f"Muestra deflactores: { {y: deflactor[y] for y in sorted(deflactor)[:5]} }", file=sys.stderr)
print(f"PIB corrientes encontrados: {len(pib_corr)} años", file=sys.stderr)
print(f"PIB constantes encontrados: {len(pib_ctes)} años", file=sys.stderr)

# ── 2. Hoja base SIIF: pivot (Nombre_Sector, Vigencia) → SUM(Valor_VF_Final) ──
# El nombre de la hoja cambia entre entregas del archivo ('BASE_SIIF_2' en
# la versión con la que se cargó la base original, 'BASE_SIIF' en la
# entrega de marzo de 2026), así que se busca entre los alias conocidos y
# se deja constancia de cuál se usó.
_HOJAS_BASE = ("BASE_SIIF_2", "BASE_SIIF")
_hoja = next((h for h in _HOJAS_BASE if h in wb.sheetnames), None)
if _hoja is None:
    sys.exit(
        f"ERROR: no se encontró ninguna hoja base de SIIF ({', '.join(_HOJAS_BASE)}).\n"
        f"Hojas disponibles: {wb.sheetnames}"
    )
print(f"Hoja base SIIF: '{_hoja}'", file=sys.stderr)
ws_src = wb[_hoja]
src_rows = list(ws_src.iter_rows(values_only=True))
wb.close()

# La fila de encabezados no siempre es la primera: la entrega de marzo de
# 2026 trae una fila en blanco arriba. Se localiza por contenido.
def _fila_encabezado(filas, requeridas=('Nombre_Sector', 'Vigencia'), maximo=10):
    for i, fila in enumerate(filas[:maximo]):
        etiquetas = {clean(c) for c in fila if c is not None}
        if all(r in etiquetas for r in requeridas):
            return i
    return None


_i_header = _fila_encabezado(src_rows)
if _i_header is None:
    sys.exit(
        "ERROR: no se encontró la fila de encabezados con 'Nombre_Sector' y "
        f"'Vigencia' en las primeras filas de '{_hoja}'."
    )
if _i_header:
    print(f"Encabezados en la fila {_i_header + 1} de '{_hoja}'", file=sys.stderr)

src_header = src_rows[_i_header]
SC = {clean(h): i for i, h in enumerate(src_header) if h is not None}

VIG_C  = SC.get('Vigencia')
VAL_C  = SC.get('Valor_VF_Final (Actual)') or SC.get('Valor_VF_Final (Actual) ')
SECT_C = SC.get('Nombre_Sector')

if None in (VIG_C, VAL_C, SECT_C):
    faltan = [n for n, c in (("Vigencia", VIG_C),
                             ("Valor_VF_Final (Actual)", VAL_C),
                             ("Nombre_Sector", SECT_C)) if c is None]
    sys.exit(
        f"ERROR: faltan columnas en la hoja '{_hoja}': {faltan}\n"
        f"Encabezados leídos: {sorted(SC)}\n\n"
        "'Valor_VF_Final (Actual)' es una columna CALCULADA que existe en la\n"
        "hoja 'BASE_SIIF_2'. Si el libro solo trae 'BASE_SIIF' con las\n"
        "columnas crudas de SIIF (ACTUAL/Autorizada/Utilizada), es una\n"
        "versión anterior del archivo: no se puede derivar aquí el valor\n"
        "final sin conocer la fórmula, y elegir una columna cruda daría\n"
        "cifras equivocadas. Solicitar el libro que incluya 'BASE_SIIF_2'."
    )

pivot = defaultdict(lambda: defaultdict(float))
filas = 0

for r in src_rows[1:]:
    vig = clean(r[VIG_C])
    try: year = int(vig)
    except: continue
    if year < 2025 or year > 2054:
        continue
    sect = clean(r[SECT_C])
    if not sect:
        continue
    val = r[VAL_C]
    if isinstance(val, (int, float)):
        pivot[sect][year] += float(val)
    else:
        s = clean(val).replace('.', '').replace(',', '.')
        try: pivot[sect][year] += float(s)
        except: pass
    filas += 1

all_sectors = sorted(pivot.keys())
all_years   = sorted({y for d in pivot.values() for y in d})

print(f"\nPivot: {filas} filas leídas → {len(all_sectors)} sectores × {len(all_years)} años",
      file=sys.stderr)

# ── 3. Cargar en BD ───────────────────────────────────────────────────────────
conn = dbmod.conectar()

bid = dbmod.bitacora_reciente(conn)
print(f"bitacora_id={bid}", file=sys.stderr)

# Este cargador es la fuente autoritativa de la sección 5: reemplaza por
# completo lo que haya, incluido lo que cargue load_bitacora_excel.py
# desde la hoja TD BITACORA, que cubre menos años y más sectores.
# Antes esto se conseguía con DROP TABLE; ahora se vacía por bitácora,
# porque borrar la tabla rompería las claves foráneas.
conn.vaciar_bitacora(("vigencias_futuras", "deflactores_pib"), bid)

# Deflactores
defl_rows = []
for y in sorted(year_col):
    d  = deflactor.get(y)
    pc = pib_corr.get(y)
    pk = pib_ctes.get(y)
    if d is not None:
        defl_rows.append((bid, y, d, pc, pk))

conn.upsert(
    "deflactores_pib",
    ["bitacora_id", "anio", "deflactor", "pib_corriente_mmm", "pib_constante_mmm"],
    defl_rows, claves=["bitacora_id", "anio"],
)

# Vigencias futuras (pivot crudo)
vf_rows = []
for sect in all_sectors:
    for year in all_years:
        pesos = pivot[sect].get(year, 0.0)
        if pesos == 0.0:
            continue
        vf_rows.append((bid, year, sect, round(pesos / 1e9, 3)))

conn.upsert(
    "vigencias_futuras",
    ["bitacora_id", "vigencia_exec", "sector", "valor_corriente_mmm"],
    vf_rows, claves=["bitacora_id", "vigencia_exec", "sector"],
)

conn.commit()

n_defl = conn.execute(
    "SELECT COUNT(*) FROM deflactores_pib WHERE bitacora_id=?", (bid,)
).fetchone()[0]
n_vf = conn.execute(
    "SELECT COUNT(*) FROM vigencias_futuras WHERE bitacora_id=?", (bid,)
).fetchone()[0]
conn.close()

print(f"\n✓ deflactores_pib:   {n_defl} filas cargadas")
print(f"✓ vigencias_futuras: {n_vf} filas cargadas  ({len(all_sectors)} sectores × años)")

# Verificación
print("\n=== DEFLACTORES (muestra) ===")
for y, d, pc, pk in [(r[1], r[2], r[3], r[4]) for r in defl_rows if r[1] in [2026,2027,2028,2030]]:
    print(f"  {y}: defl={d:.6f}  PIB_corr={pc}  PIB_ctes={pk}")

print("\n=== VF 2027 POR SECTOR (mmm corrientes) ===")
for sect in all_sectors:
    v = pivot[sect].get(2027, 0) / 1e9
    if v > 0:
        print(f"  {sect}: {v:.0f}")
