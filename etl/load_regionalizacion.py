"""
etl/load_regionalizacion.py
Carga datos de regionalización de inversión 2022-2026 desde Excel.
Fuente: BASES_BITACORA/2026/Marzo/3. REGIONALIZACIÓN/Consolidado Reg-Ejec-Marzo-2022-2026-Graficasvf.xlsx

Nota: el archivo _v_2.0 corrige una inconsistencia de datos detectada en las
columnas de 2022 y 2026 del archivo original (Consolidado Reg-Ejec-Marzo-2022-2026.xlsx,
sin sufijo de versión). Las columnas 2023-2025 no cambiaron entre versiones.
El archivo _v_2.0 fue reemplazado por "...-Graficasvf.xlsx" (misma hoja
"Regionalizacion Mar-2022-2026"); a partir de esa versión, la región INSULAR
(Archipiélago de San Andrés) se reporta consolidada dentro de CARIBE.

Uso:
    python etl/load_regionalizacion.py
    python etl/load_regionalizacion.py --xlsx "ruta/al/archivo.xlsx"
"""

import argparse
from pathlib import Path

import bases
import db as dbmod

try:
    import openpyxl
except ImportError:
    raise SystemExit("Instala openpyxl: pip install openpyxl")

# ── Rutas por defecto ──────────────────────────────────────────
DEFAULT_XLSX = bases.excel(3, "Consolidado Reg-Ejec-*Graficas*.xlsx")

# ── Mapeo nombre Excel → código DANE ──────────────────────────
DANE_MAP: dict[str, str] = {
    "Antioquia":            "05",
    "Atlántico":            "08",
    "Bogotá, D.C.":         "11",
    "Bolívar":              "13",
    "Boyacá":               "15",
    "Caldas":               "17",
    "Caquetá":              "18",
    "Cauca":                "19",
    "Cesar":                "20",
    "Córdoba":              "23",
    "Cundinamarca":         "25",
    "Chocó":                "27",
    "Huila":                "41",
    "La Guajira":           "44",
    "Magdalena":            "47",
    "Meta":                 "50",
    "Nariño":               "52",
    "Norte de Santander":   "54",
    "Quindio":              "63",   # sin tilde en el Excel
    "Quindío":              "63",
    "Risaralda":            "66",
    "Santander":            "68",
    "Sucre":                "70",
    "Tolima":               "73",
    "Valle del Cauca":      "76",
    "Arauca":               "81",
    "Casanare":             "85",
    "Putumayo":             "86",
    "Archipiélago de San Andrés, Providencia y Santa Catalina": "88",
    "Amazonas":             "91",
    "Guainía":              "94",
    "Guaviare":             "95",
    "Vaupés":               "97",
    "Vichada":              "99",
}

# ── Normalización de nombres de región ────────────────────────
# El Archipiélago de San Andrés, Providencia y Santa Catalina (región
# INSULAR) se reporta consolidado dentro de la región Caribe desde la
# Bitácora 2026-I, siguiendo la fuente oficial (Consolidado Reg-Ejec).
REGION_NORM: dict[str, str] = {
    "AMAZONAS":  "AMAZONIA",   # evita colisión con departamento homónimo
    "ANDINA":    "ANDINA",
    "CARIBE":    "CARIBE - INSULAR",
    "INSULAR":   "CARIBE - INSULAR",
    "ORINOQUÍA": "ORINOQUÍA",
    "PACÍFICO":  "PACÍFICO",
}

# Columnas (0-indexed) de cada año: (inicio_slice, fin_slice)
# Año  2022: cols 2-5  → índices 1-4
# Año  2023: cols 7-10 → índices 6-9
# Año  2024: cols 12-15 → índices 11-14
# Año  2025: cols 16-19 → índices 15-18
# Año  2026: cols 20-23 → índices 19-22
YEAR_COLS: dict[int, tuple[int, int]] = {
    2022: (1,  5),
    2023: (6,  10),
    2024: (11, 15),
    2025: (15, 19),
    2026: (19, 23),
}

SKIP_LABELS = {"Total general", "Fuente: DNP - DIFP.", "Fuente: DNP - DIFP"}


def _safe_float(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _pct(num: float, den: float) -> float | None:
    if den and den > 0:
        return round(num * 100.0 / den, 2)
    return None


def parse_sheet(ws) -> list[dict]:
    """Extrae filas normalizadas de la hoja de regionalización."""
    records = []
    current_region = None

    # Pre-calcular totales generales por año (fila 'Total general')
    totals: dict[int, float] = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        label = str(row[0]).strip() if row[0] else ""
        if label == "Total general":
            for year, (s, e) in YEAR_COLS.items():
                totals[year] = _safe_float(row[s])
            break

    # Iterar filas de datos
    for row in ws.iter_rows(min_row=6, values_only=True):
        raw_label = row[0]
        if raw_label is None:
            continue
        label = str(raw_label).strip()

        if not label or any(label.startswith(s) for s in SKIP_LABELS):
            continue

        # Detectar fila de cabecera de región (todo mayúsculas)
        if label == label.upper() and label not in ("Por Regionalizar", "Nacional"):
            norm = REGION_NORM.get(label)
            if norm:
                current_region = norm
            # Filas CAPS se omiten (se recalculan como suma de dptos)
            continue

        # Determinar tipo y región
        if label == "Por Regionalizar":
            tipo = "por_regionalizar"
            region = "POR_REGIONALIZAR"
            codigo_dane = None
        elif label == "Nacional":
            tipo = "nacional"
            region = "NACIONAL"
            codigo_dane = None
        else:
            tipo = "departamento"
            region = current_region or "SIN_REGION"
            codigo_dane = DANE_MAP.get(label)
            if codigo_dane is None:
                print(f"  [AVISO] Sin código DANE para: '{label}'")

        # Extraer valores por año
        for year, (s, e) in YEAR_COLS.items():
            vals = [_safe_float(row[i]) for i in range(s, e)]
            aprop, comp, oblig, pago = vals

            # Convertir millones → miles de millones
            aprop_mmm  = round(aprop  / 1000, 6)
            comp_mmm   = round(comp   / 1000, 6)
            oblig_mmm  = round(oblig  / 1000, 6)
            pago_mmm   = round(pago   / 1000, 6)

            total_anio = totals.get(year, 0)

            records.append({
                "vigencia":         year,
                "tipo":             tipo,
                "region":           region,
                "departamento":     label,
                "codigo_dane":      codigo_dane,
                "apropiacion_mmm":  aprop_mmm,
                "compromisos_mmm":  comp_mmm,
                "obligaciones_mmm": oblig_mmm,
                "pagos_mmm":        pago_mmm,
                "pct_compromisos":  _pct(comp,  aprop),
                "pct_obligaciones": _pct(oblig, aprop),
                "pct_pagos":        _pct(pago,  aprop),
                "pct_participacion": _pct(aprop, total_anio),
            })

    return records


def run(xlsx_path: Path) -> None:
    print(f"Leyendo: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Regionalizacion Mar-2022-2026"]

    records = parse_sheet(ws)
    print(f"Registros extraídos: {len(records)}")

    conn = dbmod.conectar()

    bid = dbmod.bitacora_reciente(conn)
    print(f"Cargando para bitacora_id={bid}")

    # Limpiar datos previos de esta bitácora
    deleted = conn.execute(
        "DELETE FROM dbo.regionalizacion WHERE bitacora_id=?", (bid,)
    ).rowcount
    if deleted > 0:
        print(f"  Eliminados {deleted} registros previos")

    # pyodbc solo admite parámetros posicionales, no :nombre como sqlite3,
    # así que los diccionarios se ordenan según COLUMNAS.
    COLUMNAS = [
        "bitacora_id", "vigencia", "tipo", "region", "departamento", "codigo_dane",
        "apropiacion_mmm", "compromisos_mmm", "obligaciones_mmm", "pagos_mmm",
        "pct_compromisos", "pct_obligaciones", "pct_pagos", "pct_participacion",
    ]
    conn.upsert(
        "regionalizacion",
        COLUMNAS,
        [tuple({**r, "bitacora_id": bid}[c] for c in COLUMNAS) for r in records],
        claves=["bitacora_id", "vigencia", "region", "departamento"],
    )

    conn.commit()

    # Resumen
    counts = conn.execute("""
        SELECT vigencia, tipo, COUNT(*) AS n
        FROM dbo.regionalizacion WHERE bitacora_id=?
        GROUP BY vigencia, tipo ORDER BY vigencia, tipo
    """, (bid,)).fetchall()
    print("\nResumen cargado:")
    for vigencia, tipo, n in counts:
        print(f"  {vigencia} - {tipo:<20} : {n} registros")

    total = conn.execute(
        "SELECT COUNT(*) FROM dbo.regionalizacion WHERE bitacora_id=?", (bid,)
    ).fetchone()[0]
    print(f"\nTotal: {total} registros en tabla 'regionalizacion'")
    conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Carga regionalización 2022-2026")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    args = parser.parse_args()
    run(args.xlsx)
