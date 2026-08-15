"""
etl/load_bitacora_excel.py -- ETL directo desde archivos Excel para nueva Bitácora PGN
Lee los archivos fuente sin necesidad de CSVs intermedios.

Secciones que carga:
  Sec 1 -- Transformaciones PND, componentes y ejecución por transformación
  Sec 4 -- Apropiación por sector (vigencia actual)
  Sec 5 -- Vigencias futuras por sector y año (TD BITACORA)
  Sec 6 -- Ejecución sectorial por entidad

Pendientes (sin archivos fuente en la primera entrega):
  Sec 2 -- Evolución presupuestal histórica
  Sec 3 -- Regionalización
  Sec 4 -- Series históricas ejecución 2022-N
  Sec 6 -- Ejecución sectorial mensual

Uso:
  python etl/load_bitacora_excel.py
  python etl/load_bitacora_excel.py --periodo 2026-I --corte 2026-03-31
  python etl/load_bitacora_excel.py --replace   (elimina y recarga si ya existe)

Requiere:  openpyxl  (pip install openpyxl)
"""

import bases
import db as dbmod
import argparse
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    raise SystemExit("Falta openpyxl. Instalar con: pip install openpyxl")

# ---------------------------------------------------------------------------
# Rutas
# ---------------------------------------------------------------------------
FILE1_DEFAULT = bases.excel(1, "Inversiones*.xlsx")
FILE5_DEFAULT = bases.excel(5, "*Base VF*.xlsx")

MMM = 1_000_000_000  # pesos -> miles de millones
TOP_COMPONENTES = 5  # cuántos componentes mostrar por transformación (resto -> "OTROS")


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------
def get_conn():
    return dbmod.conectar()


def delete_bitacora(conn, bid):
    # Orden inverso a las FK. Ya no van evolucion_presupuestal ni las
    # regionalizacion_* antiguas: se descartaron en la migración.
    tablas = [
        "inversion_transformaciones", "inversion_componentes_pnd",
        "ejecucion_transformaciones", "apropiacion_por_sector",
        "compromisos_pct_por_sector", "obligaciones_pct_por_sector",
        "pagos_pct_por_sector", "vigencias_futuras", "deflactores_pib",
        "ejecucion_sectorial_entidades", "ejecucion_sectorial_mensual",
        "ejecucion_historica", "regionalizacion", "regionalizacion_sectores",
        "credito_portafolio", "credito_ejecucion_entidad",
        "credito_ejecucion_historica",
        "sgp_historico_participacion", "sgp_historico_componentes",
    ]
    conn.vaciar_bitacora(tablas, bid)
    conn.execute("DELETE FROM dbo.metadatos_bitacora WHERE id=?", (bid,))
    conn.commit()


def create_bitacora(conn, numero, periodo, corte, fuente, notas):
    existing = conn.execute(
        "SELECT id FROM metadatos_bitacora WHERE periodo=?", (periodo,)
    ).fetchone()
    if existing:
        return None, existing[0]
    nuevo_id = conn.insertar_devolviendo_id(
        """INSERT INTO dbo.metadatos_bitacora
               (numero_bitacora, periodo, corte_fecha, fuente_principal, notas)
           VALUES (?,?,?,?,?)""",
        (numero, periodo, corte, fuente, notas),
    )
    conn.commit()
    return nuevo_id, None


# ---------------------------------------------------------------------------
# Lectura y agregación de la hoja Base (una sola pasada)
# ---------------------------------------------------------------------------
def read_base_sheet(ws):
    """
    Itera la hoja Base una sola vez y devuelve datos agregados para
    Sec 1 (transformaciones/componentes), Sec 4 (sectores) y Sec 6 (entidades).

    Columnas (0-indexed):
      1=Sector, 3=Unidad Ejecutora, 17=Vigente, 19=Compromisos,
      20=Obligación, 21=Pago, 24=Transformación, 25=Componente
    """
    # Acumuladores
    t_acc   = defaultdict(lambda: {"v": 0.0, "c": 0.0, "o": 0.0, "p": 0.0})
    tc_acc  = defaultdict(lambda: defaultdict(float))  # t -> comp -> vigente
    s_acc   = defaultdict(lambda: {"v": 0.0, "c": 0.0, "o": 0.0, "p": 0.0})
    ent_acc = defaultdict(lambda: {"sector": "", "v": 0.0, "c": 0.0, "o": 0.0})

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row[0]:   # encabezado o fila vacía
            continue

        sector  = str(row[1]).strip()  if row[1]  else "SIN SECTOR"
        entidad = str(row[3]).strip()  if row[3]  else sector
        t       = str(row[24]).strip() if row[24] else None
        comp    = str(row[25]).strip() if row[25] else None

        v  = float(row[17] or 0)
        c  = float(row[19] or 0)
        o  = float(row[20] or 0)
        p  = float(row[21] or 0)

        # Transformaciones
        if t:
            t_acc[t]["v"] += v
            t_acc[t]["c"] += c
            t_acc[t]["o"] += o
            t_acc[t]["p"] += p
            if comp:
                tc_acc[t][comp] += v

        # Sectores
        s_acc[sector]["v"] += v
        s_acc[sector]["c"] += c
        s_acc[sector]["o"] += o
        s_acc[sector]["p"] += p

        # Entidades
        ent_acc[entidad]["sector"] = sector
        ent_acc[entidad]["v"] += v
        ent_acc[entidad]["c"] += c
        ent_acc[entidad]["o"] += o

    return t_acc, tc_acc, s_acc, ent_acc


# ---------------------------------------------------------------------------
# Carga Sección 1
# ---------------------------------------------------------------------------
def load_sec1(conn, bid, vigencia, t_acc, tc_acc):
    total_v = sum(d["v"] for d in t_acc.values()) or 1

    # inversion_transformaciones
    rows = [
        (bid, vigencia, t,
         round(d["v"] / MMM, 3),
         round(d["v"] / total_v * 100, 2))
        for t, d in t_acc.items()
    ]
    conn.upsert(
        "inversion_transformaciones",
        ["bitacora_id", "vigencia", "transformador", "inversion_mmm", "peso_pct"],
        rows, claves=["bitacora_id", "vigencia", "transformador"],
    )
    print(f"  [OK] inversion_transformaciones      : {len(rows):>4} filas")

    # inversion_componentes_pnd  (top N + "OTROS COMPONENTES")
    rows_c = []
    for t, comps in tc_acc.items():
        t_total = t_acc[t]["v"] or 1
        sorted_c = sorted(comps.items(), key=lambda x: -x[1])
        top = sorted_c[:TOP_COMPONENTES]
        otros_v = sum(v for _, v in sorted_c[TOP_COMPONENTES:])

        for comp, v in top:
            rows_c.append((bid, vigencia, t, comp,
                           round(v / MMM, 3),
                           round(v / t_total * 100, 2)))
        if otros_v > 0:
            rows_c.append((bid, vigencia, t, "OTROS COMPONENTES",
                           round(otros_v / MMM, 3),
                           round(otros_v / t_total * 100, 2)))

    conn.upsert(
        "inversion_componentes_pnd",
        ["bitacora_id", "vigencia", "transformador", "componente", "vigente_mmm", "peso_pct"],
        rows_c, claves=["bitacora_id", "vigencia", "transformador", "componente"],
    )
    print(f"  [OK] inversion_componentes_pnd       : {len(rows_c):>4} filas")

    # ejecucion_transformaciones
    rows_e = []
    for t, d in t_acc.items():
        v = d["v"] or 1
        rows_e.append((
            bid, vigencia, t,
            round(d["v"] / MMM, 3),
            round(d["c"] / MMM, 3),
            round(d["o"] / MMM, 3),
            round(d["p"] / MMM, 3),
            round(d["c"] / v * 100, 1),
            round(d["o"] / v * 100, 1),
            round(d["p"] / v * 100, 1),
        ))
    conn.upsert(
        "ejecucion_transformaciones",
        ["bitacora_id", "vigencia", "transformador", "apr_vigente_mmm",
         "compromisos_mmm", "obligaciones_mmm", "pagos_mmm",
         "pct_c_av", "pct_o_av", "pct_p_av"],
        rows_e, claves=["bitacora_id", "vigencia", "transformador"],
    )
    print(f"  [OK] ejecucion_transformaciones      : {len(rows_e):>4} filas")


# ---------------------------------------------------------------------------
# Carga Sección 4 (apropiación sectorial) y Sección 6 (entidades)
# ---------------------------------------------------------------------------
def load_sec4_sec6(conn, bid, vigencia, s_acc, ent_acc):
    # apropiacion_por_sector
    rows_ap = [
        (bid, vigencia, s, round(d["v"] / MMM, 3))
        for s, d in s_acc.items()
    ]
    conn.upsert(
        "apropiacion_por_sector",
        ["bitacora_id", "vigencia", "sector", "vigente_mmm"],
        rows_ap, claves=["bitacora_id", "vigencia", "sector"],
    )
    print(f"  [OK] apropiacion_por_sector          : {len(rows_ap):>4} filas")

    # compromisos / obligaciones / pagos pct por sector
    for tabla, campo, clave in [
        ("compromisos_pct_por_sector",  "pct_compromisos",  "c"),
        ("obligaciones_pct_por_sector", "pct_obligaciones", "o"),
        ("pagos_pct_por_sector",        "pct_pagos",        "p"),
    ]:
        rows_pct = [
            (bid, vigencia, s,
             round(d[clave] / d["v"] * 100, 1) if d["v"] else None)
            for s, d in s_acc.items()
        ]
        conn.upsert(
            tabla,
            ["bitacora_id", "vigencia", "sector", campo],
            rows_pct, claves=["bitacora_id", "vigencia", "sector"],
        )
    print(f"  [OK] compromisos/obligaciones/pagos_pct_por_sector: {len(rows_ap):>4} filas c/u")

    # ejecucion_sectorial_entidades
    # Fila de total por sector
    rows_ent = []
    for s, d in s_acc.items():
        v = d["v"] or 1
        rows_ent.append((
            bid, vigencia, s.upper(), f"SECTOR {s.upper()}",
            round(d["v"] / MMM, 3),
            round(d["c"] / MMM, 3),
            round(d["o"] / MMM, 3),
            round(d["c"] / v * 100, 1),
            round(d["o"] / v * 100, 1),
        ))
    # Fila por cada entidad
    for ent, d in ent_acc.items():
        v = d["v"] or 1
        sector_name = d["sector"].upper()
        # Evitar duplicar la fila de sector total con nombre de entidad igual al sector
        if ent.upper() == sector_name:
            continue
        rows_ent.append((
            bid, vigencia, sector_name, ent,
            round(d["v"] / MMM, 3),
            round(d["c"] / MMM, 3),
            round(d["o"] / MMM, 3),
            round(d["c"] / v * 100, 1),
            round(d["o"] / v * 100, 1),
        ))
    conn.upsert(
        "ejecucion_sectorial_entidades",
        ["bitacora_id", "vigencia", "sector", "entidad", "apr_vigente_mmm",
         "compromisos_mmm", "obligaciones_mmm", "pct_c_av", "pct_o_av"],
        rows_ent, claves=["bitacora_id", "vigencia", "entidad"],
    )
    n_sectores = len(s_acc)
    n_entidades = len(rows_ent) - n_sectores
    print(f"  [OK] ejecucion_sectorial_entidades   : {len(rows_ent):>4} filas  ({n_sectores} sect. + {n_entidades} entid.)")


# ---------------------------------------------------------------------------
# Carga Sección 5 -- Vigencias Futuras (hoja TD BITACORA)
# ---------------------------------------------------------------------------
def load_sec5(conn, bid, ws_vf, año_inicio, año_fin):
    """
    Estructura de TD BITACORA:
      Fila 7 (idx 6): encabezados -- col 0='Etiquetas de fila', col 1=2025, col 2=2026, ...
      Fila 8+:        datos -- col 0=Sector, cols 1-N=valor pesos corrientes
    """
    rows_iter = ws_vf.iter_rows(values_only=True)

    # Saltar hasta la fila de encabezados (índice 6 -> 7ª fila)
    header = None
    for idx, row in enumerate(rows_iter):
        if idx == 6:
            header = list(row)
            break

    if header is None:
        print("  [!]  No se encontró fila de encabezados (índice 6) en TD BITACORA")
        return

    # Mapear columna -> año
    col_año = {
        j: int(v)
        for j, v in enumerate(header)
        if isinstance(v, (int, float)) and año_inicio <= int(v) <= año_fin
    }
    if not col_año:
        print(f"  [!]  No hay columnas para años {año_inicio}-{año_fin} en TD BITACORA")
        return

    rows_vf = []
    skip_labels = {"total general", "etiquetas de fila"}
    for row in rows_iter:
        r = list(row)
        sector = str(r[0]).strip() if r[0] else None
        if not sector or sector.lower() in skip_labels:
            continue
        for j, año in col_año.items():
            val = r[j] if j < len(r) else None
            if val and isinstance(val, (int, float)) and val > 0:
                rows_vf.append((bid, año, sector.upper(), round(val / MMM, 3), None))

    conn.upsert(
        "vigencias_futuras",
        ["bitacora_id", "vigencia_exec", "sector", "valor_corriente_mmm"],
        [(b, a, sec, v) for b, a, sec, v, _ in rows_vf],
        claves=["bitacora_id", "vigencia_exec", "sector"],
    )
    print(f"  [OK] vigencias_futuras               : {len(rows_vf):>4} filas  (anos {año_inicio}-{año_fin})")
    print("       [!] Valores en pesos corrientes. Deflactación aplicada en API (/api/vigencias_futuras/chart).")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Carga nueva Bitácora PGN desde archivos Excel fuente"
    )
    parser.add_argument("--numero",    default="3",            help="Número de bitácora (default: 3)")
    parser.add_argument("--periodo",   default="2026-I",       help="Período, ej. 2026-I (default: 2026-I)")
    parser.add_argument("--corte",     default="2026-03-31",   help="Fecha de corte YYYY-MM-DD (default: 2026-03-31)")
    parser.add_argument("--vigencia",  type=int, default=2026, help="Año de la vigencia presupuestal (default: 2026)")
    parser.add_argument("--fuente",    default="SIIF Nación / DPIP", help="Fuente de datos")
    parser.add_argument("--notas",     default="PND 2022-2026 Colombia potencia mundial de la vida. Corte 31 de marzo de 2026.")
    parser.add_argument("--file1",     default=str(FILE1_DEFAULT), help="Excel de Inversiones PND (Sec 1)")
    parser.add_argument("--file5",     default=str(FILE5_DEFAULT), help="Excel de Vigencias Futuras (Sec 5)")
    parser.add_argument("--vf-inicio", type=int, default=2026, help="Primer año VF a cargar (default: 2026)")
    parser.add_argument("--vf-fin",    type=int, default=2040, help="Último año VF a cargar (default: 2040)")
    parser.add_argument("--replace",   action="store_true",    help="Eliminar bitácora existente y recargar")
    args = parser.parse_args()

    conn = get_conn()

    # Manejo de bitácora existente
    existing = conn.execute(
        "SELECT id FROM metadatos_bitacora WHERE periodo=?", (args.periodo,)
    ).fetchone()

    if existing:
        if args.replace:
            bid_old = existing[0]
            delete_bitacora(conn, bid_old)
            print(f"[DEL]   Bitácora {args.periodo} (id={bid_old}) eliminada para recarga")
        else:
            print(f"[!]   Ya existe la bitácora '{args.periodo}' (id={existing[0]}).")
            print("    Use --replace para eliminarla y recargar desde cero.")
            conn.close()
            return

    bid, _ = create_bitacora(conn, args.numero, args.periodo, args.corte, args.fuente, args.notas)
    print(f"\n[OK]  Bitácora creada: {args.periodo}  (id={bid}, corte={args.corte})")

    # ---- Archivo 1 -- Inversiones PND ----
    f1 = Path(args.file1)
    if f1.exists():
        print(f"\n[DIR]  Archivo 1: {f1.name}")
        wb1 = openpyxl.load_workbook(str(f1), read_only=True, data_only=True)
        if "Base" in wb1.sheetnames:
            print("  Leyendo hoja Base (una pasada)...")
            t_acc, tc_acc, s_acc, ent_acc = read_base_sheet(wb1["Base"])
            print(f"  -- {sum(1 for _ in t_acc)} transformaciones, {len(s_acc)} sectores, {len(ent_acc)} entidades")
            print("  Cargando Sección 1 (Transformaciones PND)...")
            load_sec1(conn, bid, args.vigencia, t_acc, tc_acc)
            print("  Cargando Sección 4 (Apropiación sectorial) + Sección 6 (Entidades)...")
            load_sec4_sec6(conn, bid, args.vigencia, s_acc, ent_acc)
        else:
            print("  [!]  No se encontró la hoja 'Base' en el archivo 1")
    else:
        print(f"\n[!]  Archivo 1 no encontrado: {f1}")
        print("    Sec 1, 4 (parcial) y 6 no se cargaron.")

    # ---- Archivo 5 -- Vigencias Futuras ----
    f5 = Path(args.file5)
    if f5.exists():
        print(f"\n[DIR]  Archivo 5: {f5.name}")
        wb5 = openpyxl.load_workbook(str(f5), read_only=True, data_only=True)
        if "TD BITACORA" in wb5.sheetnames:
            print("  Cargando Sección 5 (Vigencias Futuras)...")
            load_sec5(conn, bid, wb5["TD BITACORA"], args.vf_inicio, args.vf_fin)
        else:
            print("  [!]  No se encontró la hoja 'TD BITACORA' en el archivo 5")
    else:
        print(f"\n[!]  Archivo 5 no encontrado: {f5}")
        print("    Sección 5 no se cargó.")

    conn.commit()

    # Resumen final
    print(f"\n{'-'*55}")
    print(f"  Bitácora {args.periodo} (id={bid}) -- Resumen")
    print(f"{'-'*55}")
    tablas_check = [
        "inversion_transformaciones",
        "inversion_componentes_pnd",
        "ejecucion_transformaciones",
        "apropiacion_por_sector",
        "compromisos_pct_por_sector",
        "vigencias_futuras",
        "ejecucion_sectorial_entidades",
    ]
    for tbl in tablas_check:
        try:
            n = conn.execute(
                f"SELECT COUNT(*) FROM {tbl} WHERE bitacora_id=?", (bid,)
            ).fetchone()[0]
            estado = "[OK]" if n > 0 else "[ ]"
            print(f"  {estado}  {tbl:<40} {n:>5} filas")
        except Exception as e:
            print(f"  [X]  {tbl:<40} error: {e}")

    pendientes = [
        ("Sec 2", "evolucion_presupuestal",        "carpeta 2. EVOLUCIÓN PRESUPUESTAL"),
        ("Sec 3", "regionalizacion_detalle_2025",  "carpeta 3. REGIONALIZACIÓN"),
        ("Sec 4", "ejecucion_historica",            "carpeta 4. EJECUCIÓN (series 2022-N)"),
        ("Sec 6", "ejecucion_sectorial_mensual",   "carpeta 6. EJECUCIÓN SECTORIAL"),
    ]
    print(f"\n  Pendientes (archivos fuente no disponibles):")
    for sec, tbl, origen in pendientes:
        print(f"  [ ]  {sec}  {tbl:<40} <- {origen}")

    conn.close()
    print(f"\n[DONE]  Proceso finalizado.\n")


if __name__ == "__main__":
    main()
