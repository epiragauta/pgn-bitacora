"""
ETL: carga hoja sectores_por_region del Excel de regionalización.
Valores en pesos → divide por 1_000_000_000 para obtener mmm.

Nota: sigue apuntando al Consolidado sin sufijo de versión (no al _v_2.0)
porque la hoja "sectores_por_region" no existe en Consolidado Reg-Ejec-
Marzo-2022-2026_v_2.0.xlsx. La corrección de datos que motivó la v_2.0
solo afectó la hoja "Regionalizacion Mar-2022-2026" (ver load_regionalizacion.py).
"""
import openpyxl

import bases
import db as dbmod

# El nombre del archivo varía entre entregas ('...2022-2026.xlsx',
# '...2022-2026vf.xlsx'); se excluye el de gráficas, que usa otro cargador.
EXCEL = next(
    f for f in [bases.carpeta_seccion(3) / n for n in (
        'Consolidado Reg-Ejec-Marzo-2022-2026.xlsx',
        'Consolidado Reg-Ejec-Marzo-2022-2026vf.xlsx',
    )] if f.exists()
)

REGION_NORM = {
    'andina':    'ANDINA',
    'caribe':    'CARIBE - INSULAR',
    'pacifico':  'PACÍFICO',
    'pacífico':  'PACÍFICO',
    'orinoquia': 'ORINOQUÍA',
    'orinoquía': 'ORINOQUÍA',
    'amazonas':  'AMAZONIA',
    'amazonia':  'AMAZONIA',
    'insular':   'CARIBE - INSULAR',
}

def normalize_region(name):
    if not name:
        return None
    return REGION_NORM.get(name.strip().lower())

def run():
    # El esquema de regionalizacion_sectores vive en db/mssql/001_schema.sql;
    # este cargador ya no lo crea.
    conn = dbmod.conectar()

    bid = dbmod.bitacora_reciente(conn)
    print(f"Bitácora activa: id={bid}")

    wb = openpyxl.load_workbook(EXCEL, data_only=True)

    HOJA = 'sectores_por_region'
    if HOJA not in wb.sheetnames:
        raise SystemExit(
            f"ERROR: la hoja '{HOJA}' no existe en {EXCEL.name}.\n"
            f"Hojas disponibles: {wb.sheetnames}\n\n"
            "Esa hoja es una tabla ya consolidada (región, sector, apropiación,\n"
            "compromisos, obligaciones, pagos, vigencia). El libro entregado solo\n"
            "trae las hojas por región sin consolidar, así que derivarla aquí\n"
            "significaría reimplementar una agregación que no está documentada y\n"
            "arriesgar cifras erróneas. Solicitar el libro que incluya la hoja."
        )
    ws = wb[HOJA]

    VIGENCIA_MIN, VIGENCIA_MAX = 2000, 2100

    filas = []
    skipped = 0
    vigencias_invalidas = []
    for n_fila, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        region_raw, sector, aprop, comp, obl, pag, vigencia = row
        if not region_raw or not sector or not vigencia:
            continue

        region = normalize_region(region_raw)
        if not region:
            print(f"  WARN: región desconocida '{region_raw}'")
            skipped += 1
            continue

        # Un año fuera de rango es casi siempre una celda mal digitada. Cargarlo
        # no daría error —la columna admite cualquier entero— pero dejaría la
        # fila fuera del año consultado por la API, así que ese sector
        # desaparecería del tablero sin ninguna señal.
        try:
            anio = int(vigencia)
        except (TypeError, ValueError):
            anio = None
        if anio is None or not (VIGENCIA_MIN <= anio <= VIGENCIA_MAX):
            vigencias_invalidas.append((n_fila, region_raw, str(sector).strip(), vigencia))
            continue

        filas.append((
            bid, anio, region, sector.strip(),
            round((aprop or 0) / 1e9, 3),
            round((comp  or 0) / 1e9, 3),
            round((obl   or 0) / 1e9, 3),
            round((pag   or 0) / 1e9, 3),
        ))

    if vigencias_invalidas:
        detalle = "\n".join(
            f"    fila {n} (celda G{n}): {reg} / {sec} -> Año = {v!r}"
            for n, reg, sec, v in vigencias_invalidas
        )
        raise SystemExit(
            f"ERROR: {len(vigencias_invalidas)} fila(s) de '{HOJA}' con un año "
            f"fuera del rango {VIGENCIA_MIN}-{VIGENCIA_MAX}:\n{detalle}\n\n"
            "Corregir la celda en el Excel y volver a ejecutar. No se carga nada:\n"
            "una vigencia equivocada saca a ese sector del año que consulta la\n"
            "API y desaparecería del tablero sin ningún error visible."
        )

    # El ON CONFLICT DO UPDATE de SQLite equivale al upsert por clave natural.
    n = conn.upsert(
        "regionalizacion_sectores",
        ["bitacora_id", "vigencia", "region", "sector",
         "apropiacion_mmm", "compromisos_mmm", "obligaciones_mmm", "pagos_mmm"],
        filas, claves=["bitacora_id", "vigencia", "region", "sector"],
    )
    conn.commit()
    print(f"Listo: {n} registros insertados/actualizados, {skipped} omitidos.")

    print("\n--- Verificación: top 5 sectores ANDINA 2026 ---")
    for r in conn.execute("""
        SELECT TOP 5 sector, apropiacion_mmm, compromisos_mmm
        FROM dbo.regionalizacion_sectores
        WHERE bitacora_id=? AND region='ANDINA' AND vigencia=2026
        ORDER BY apropiacion_mmm DESC
    """, (bid,)):
        print(f"  {r[0]:<45} aprop={r[1]:>8.1f} mmm  comp={r[2]:>8.1f} mmm")
    conn.close()

if __name__ == '__main__':
    run()
